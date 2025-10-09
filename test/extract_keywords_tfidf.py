from sklearn.feature_extraction.text import TfidfVectorizer
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# ---------- AM ----------
segments = open("am").read().strip().split("\n")

vectorizer = TfidfVectorizer(stop_words='english', ngram_range=(1, 2))
tfidf_matrix = vectorizer.fit_transform(segments)

feature_names = vectorizer.get_feature_names_out()
avg_tfidf = tfidf_matrix.mean(axis=0).A1

df_keywords_am = pd.DataFrame({'keyword': feature_names, 'score': avg_tfidf})

# Filter: keep only keywords with length >= 5
df_keywords_am = df_keywords_am[df_keywords_am['keyword'].str.len() >= 5]

# Sort and keep top 20
df_keywords_am = df_keywords_am.sort_values(by='score', ascending=False).head(20)

# ---------- LIVE ----------
segments = open("live").read().strip().split("\n")

vectorizer = TfidfVectorizer(stop_words='english', ngram_range=(1, 2))
tfidf_matrix = vectorizer.fit_transform(segments)

feature_names = vectorizer.get_feature_names_out()
avg_tfidf = tfidf_matrix.mean(axis=0).A1

df_keywords_live = pd.DataFrame({'keyword': feature_names, 'score': avg_tfidf})

# Filter: keep only keywords with length >= 5
df_keywords_live = df_keywords_live[df_keywords_live['keyword'].str.len() >= 5]

# Sort and keep top 20
df_keywords_live = df_keywords_live.sort_values(by='score', ascending=False).head(20)

# ---------- MERGE ---------
df_am_filtered = df_keywords_am[(df_keywords_am['score'] > 0.05) & (~df_keywords_am['keyword'].isin(df_keywords_live['keyword']))]

blank_col = pd.DataFrame({'': []})


# Combine horizontally with a separating column
combined = pd.concat(
    [df_keywords_am.reset_index(drop=True), blank_col.reset_index(drop=True), df_keywords_live.reset_index(drop=True), blank_col.reset_index(drop=True), df_am_filtered.reset_index(drop=True)],
    axis=1, ignore_index=False
)

# Write to Excel
with pd.ExcelWriter("tfidf.xlsx", engine='openpyxl') as writer:
    combined.to_excel(writer, index=False, sheet_name="TF-IDF Combined", startrow=1)

# Add titles manually at the top
wb = load_workbook("tfidf.xlsx")
ws = wb["TF-IDF Combined"]

# Titles row
ws["A1"] = "AM tf-idf"
ws["D1"] = "Non-AM tf-idf"
ws["G1"] = "AM tf-idf (>0.05) Except Non-AM"

# Make titles bold
for cell in ["A1", "D1", "G1"]:
    ws[cell].font = Font(bold=True)

wb.save("tfidf.xlsx")
print("✅ tfidf.xlsx created with all tables side-by-side and labeled.")

